from openpyxl import load_workbook
from openpyxl.chart import Reference, BarChart, BarChart3D


def create_bar_chart_1(file_path):
    """
    插入柱形图
    :param file_path: Excel 文件路径
    :return: None
    """
    wb = load_workbook(file_path)
    st = wb.active

    data1 = Reference(st, min_col=2, min_row=1, max_row=7, max_col=3)
    cats1 = Reference(st, min_col=1, min_row=2, max_row=7)

    chart1 = BarChart()
    chart1.type = "col"
    chart1.title = "日均值对比"
    # chart1.y_axis.title = '数值'
    chart1.x_axis.title = st.cell(column=1, row=1).value

    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    st.add_chart(chart1, 'A8')
    wb.save(file_path)


def create_bar_chart_2(file_path):
    """
    插入3D柱形图
    :param file_path: Excel 文件路径
    :return: None
    """
    wb = load_workbook(file_path)
    st = wb.active

    data1 = Reference(st, min_col=2, min_row=1, max_row=7, max_col=3)
    cats1 = Reference(st, min_col=1, min_row=2, max_row=7)

    chart1 = BarChart3D()
    chart1.type = "bar"
    chart1.style = 10
    chart1.title = "日均值对比"
    chart1.y_axis.title = 'Draw by RustFisher'
    chart1.shape = 'cylinder'

    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    st.add_chart(chart1, 'A26')
    wb.save(file_path)


def create_bar_chart_for_ratios(file_path, bar_chart, chart_location='F8'):
    """
    插入柱形图
    :param file_path: Excel 文件路径
    :param bar_chart 要添加的图表
    :param chart_location 图表位置
    :return: None
    """
    wb = load_workbook(file_path)
    st = wb.active

    data1 = Reference(st, min_col=4, min_row=1, max_col=7, max_row=6)
    cats1 = Reference(st, min_col=1, min_row=2, max_row=6)

    chart1 = bar_chart
    chart1.type = "col"
    chart1.style = 10
    chart1.title = 'rustfisher.com'
    chart1.x_axis.title = None

    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    st.add_chart(chart1, chart_location)
    wb.save(file_path)


if __name__ == '__main__':
    excel_path = 'out/指数.xlsx'
    create_bar_chart_1(excel_path)
    create_bar_chart_2(excel_path)
    create_bar_chart_for_ratios(excel_path, BarChart())
    create_bar_chart_for_ratios(excel_path, BarChart3D(), chart_location='F26')
