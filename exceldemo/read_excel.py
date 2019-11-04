from copy import copy

from openpyxl import load_workbook


def read_xlsx_basic(file_path):
    """
    读取Excel的数据并打印出来
    """
    wb = load_workbook(file_path)
    st = wb.active
    end_row = st.max_row + 1
    end_column = st.max_column + 1
    print(st.title, '有', end_row, '行', end_column, '列')
    for row in range(1, end_row):
        for col in range(1, end_column):
            print('{:10}'.format(st.cell(row=row, column=col).value), end='')
        print()


def change_title_line_format(file_path):
    wb = load_workbook(file_path)
    st = wb.active

    # 调整宽度
    for col in range(ord('A'), ord('G') + 1):
        st.column_dimensions[chr(col)].width = 15

    # 调整第一行的单元格
    for col in range(1, st.max_column + 1):
        cell = st.cell(row=1, column=col)
        font = copy(cell.font)  # 调整字体
        font.size = 13
        font.bold = True
        cell.font = font
    wb.save(file_path)


if __name__ == '__main__':
    excel_path = 'out/指数.xlsx'
    ws1 = load_workbook(excel_path).active
    if ws1 is None:
        print('表格不存在')
    else:
        print('读取基本数据')
        read_xlsx_basic(excel_path)
        change_title_line_format(excel_path)
