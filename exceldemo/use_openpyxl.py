# Read write excel files.
# https://openpyxl.readthedocs.io/en/stable/index.html#module-openpyxl
import datetime
import os

from openpyxl import Workbook


def create_xlsx(file_path):
    """
    Create a new excel file.
    :param file_path: new file.
    """
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 42
    ws['B1'] = 'Days'
    ws.append([1, 2, 3, 4])
    ws.append([5, 6, 7, 8])
    ws['A2'] = datetime.datetime.now()
    wb.save(file_path)


def read_xlsx(file_path):
    """
    Read data from excel.
    """
    if not os.path.exists(file_path):
        return


if __name__ == '__main__':
    print('use openpyxl demo')
    out_dir = 'out'
    if not os.path.exists(out_dir):
        os.mkdir(out_dir)
    create_xlsx(out_dir + '/write1.xlsx')
