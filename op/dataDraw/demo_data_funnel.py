from openpyxl import load_workbook
from pyecharts.charts import Funnel
import pyecharts.options as opts


def read_in_data(file_path):
    print('获取原始数据 >>>')
    wb = load_workbook(file_path)
    st = wb['data3']
    data_list = []
    for row_index in range(2, st.max_row + 1):
        event_name = st.cell(row=row_index, column=1).value
        event_value = st.cell(row=row_index, column=2).value
        data_list.append([event_name, event_value])
    print(data_list)
    return data_list


def draw_funnel_default(input_data):
    print('绘制漏斗图 >>> https://gallery.pyecharts.org/#/Funnel/funnel_base')
    (
        Funnel(init_opts=opts.InitOpts(width="800px", height="400px"))
            .add(series_name='漏斗1', data_pair=input_data, )
            .set_global_opts(title_opts=opts.TitleOpts(title="漏斗图1", subtitle="默认样式"))
            .render("funnel_chart_1.html")
    )


def draw_funnel_2(input_data):
    print('绘制漏斗图1 >>> https://gallery.pyecharts.org/#/Funnel/funnel_base')
    (
        Funnel(init_opts=opts.InitOpts(width="1000px", height="600px"))
            .add(series_name='漏斗1', data_pair=input_data,
                 label_opts=opts.LabelOpts(position="inside", formatter='{b} 数量: {c}'),
                 gap=2,
                 tooltip_opts=opts.TooltipOpts(trigger="item", formatter="{a} <br/>{b} 数量: {c}"),
                 )
            .set_global_opts(title_opts=opts.TitleOpts(title="漏斗图1", subtitle="改变了label"))
            .render("funnel_chart_2.html")
    )


def draw_funnel_3(input_data):
    print('绘制漏斗图2 >>> https://gallery.pyecharts.org/#/Funnel/funnel_base')
    data = input_data.copy()  # 复制一份出来处理标题
    base_value = data[0][1]
    for x in range(0, len(data)):
        event_name = data[x][0]
        value = data[x][1]  # 拿到数值
        ratio = value / base_value
        event_name += '\n总体转化率: {:.2f}%'.format(ratio * 100)
        if x > 0:
            event_name += '\n相对转化率: {:.2f}%'.format(100 * value / data[x - 1][1])
        data[x][0] = event_name
        print(event_name)

    (
        Funnel(init_opts=opts.InitOpts(width="1000px", height="800px"))
            .add(series_name='漏斗1', data_pair=data,
                 label_opts=opts.LabelOpts(position="inside", formatter='{b}\n数量: {c}'),
                 gap=2,
                 tooltip_opts=opts.TooltipOpts(trigger="item", formatter="{a} <br/>{b} 数量: {c}"),
                 )
            .set_global_opts(title_opts=opts.TitleOpts(title="漏斗图1", subtitle="修改了各项的标题"))
            .render("funnel_chart_3.html")
    )


if __name__ == '__main__':
    src_data = read_in_data('res/事件demo.xlsx')
    draw_funnel_default(src_data)
    draw_funnel_2(src_data)
    draw_funnel_3(src_data)
